import pandas as pd
from openpyxl import load_workbook
import os
import re
from datetime import datetime, timedelta

# Function to find the latest week folder based on a specific pattern
def find_latest_week_folder(base_path):
    folders = [f for f in os.listdir(base_path) if os.path.isdir(os.path.join(base_path, f))]
    week_folders = [f for f in folders if re.match(r'^Week_\d+_\(\d{2}\.\d{2}\.\d{2}\)_\(\d{2}\.\d{2}\.\d{2}\)$', f)]
    
    if not week_folders:
        raise FileNotFoundError("No week folders found in the specified directory.")
    
    latest_folder = max(week_folders, key=lambda f: os.path.getmtime(os.path.join(base_path, f)))
    return os.path.join(base_path, latest_folder)

# Set base path for the dynamic search
base_path = r"C:\Users\User\Desktop\JEAV\E-Commerce Invoice (tuesday)\Curl Mix"

# Get the latest week folder dynamically
latest_week_folder = find_latest_week_folder(base_path)
print(f"Latest week folder found: {latest_week_folder}")

# Set dynamic input file path for the invoice
invoice_file_path = os.path.join(latest_week_folder, "2.Cleaning", "Invoice_(DoNotSend)_(11.24.24)_(11.30.24).xlsx")
modified_invoice_path = os.path.join(latest_week_folder, "2.Cleaning", "Modified_Invoice_(11.24.24)_(11.30.24).xlsx")

# Load the second report dynamically
invoice_folder_path = os.path.join(latest_week_folder, "3.Invoice")
# Find the invoice file for the current week
invoice_files = [f for f in os.listdir(invoice_folder_path) if re.match(r'^CM_Inv_\d+_Week_\(\d{2}\.\d{2}\.\d{2}\)_\(\d{2}\.\d{2}\.\d{2}\)\.xlsx$', f)]

if not invoice_files:
    raise FileNotFoundError("No invoice files found in the '3.Invoice' folder.")

# Assuming you want the most recent file based on filename pattern
latest_invoice_file = max(invoice_files, key=lambda f: int(re.search(r'(\d+)', f).group(1)))
second_report_path = os.path.join(invoice_folder_path, latest_invoice_file)
print(f"Using second report: {second_report_path}")

# Load the second report
second_report_df = pd.read_excel(second_report_path)

# Exclude the last row if it's the totals row
if not second_report_df.empty:
    second_report_df = second_report_df[:-1]  # Exclude the last row (totals row)

# Convert relevant columns to numeric, handling non-numeric gracefully
numeric_columns = [
    'Package Charge', 'Fuel Charge', 'Delivery Confirmation Charge',
    'OSM DIM Surcharge', 'Relabel Fee', 'Delivery Area Surcharge (DAS)', 
    'OCR Fee', 'Peak Season Surcharge', 'Nonstandard Length Fee 22 in', 
    'Nonstandard Length Fee 30 in', 'Nonstandard Length Fee 2 cu', 
    'Dimension Noncompliance', 'Irregular Shape Charge', 
    'Non Compliance/Unmanifested Charge', 'Signature Confirmation'
]

for column in numeric_columns:
    second_report_df[column] = pd.to_numeric(second_report_df[column], errors='coerce')

# Proceed with the calculations for counts and sums, excluding zero values
# Count and sum for "Parcel"
parcel_df = second_report_df[second_report_df['Billed Service'] == "Parcel"]
parcel_df = parcel_df[parcel_df['Package Charge'] > 0]  # Exclude rows where Package Charge is 0
parcel_count = parcel_df.shape[0]  # Count of rows
parcel_package_charge_sum = parcel_df['Package Charge'].sum()  # Sum of package charges

# Count and sum for "Small Parcel"
small_parcel_df = second_report_df[second_report_df['Billed Service'] == "Small Parcel"]
small_parcel_df = small_parcel_df[small_parcel_df['Package Charge'] > 0]  # Exclude rows where Package Charge is 0
small_parcel_count = small_parcel_df.shape[0]  # Count of rows
small_parcel_package_charge_sum = small_parcel_df['Package Charge'].sum()  # Sum of package charges

# Count and sum for "Ground Adv Over 1 lb"
ground_adv_df = second_report_df[second_report_df['Billed Service'] == "Ground Adv Over 1 lb"]
ground_adv_df = ground_adv_df[ground_adv_df['Package Charge'] > 0]  # Exclude rows where Package Charge is 0
ground_adv_count = ground_adv_df.shape[0]  # Count of rows
ground_adv_package_charge_sum = ground_adv_df['Package Charge'].sum()  # Sum of package charges

# Static value and sum of Fuel Charge
fuel_charge_sum = second_report_df['Fuel Charge'].sum()

# Count and sum for Delivery Confirmation Charge
delivery_confirmation_df = second_report_df[second_report_df['Delivery Confirmation Charge'] > 0]
delivery_confirmation_count = delivery_confirmation_df.shape[0]  # Count of non-null rows
delivery_confirmation_sum = delivery_confirmation_df['Delivery Confirmation Charge'].sum()  # Sum

# Count and sum for OSM DIM Surcharge
osm_dim_df = second_report_df[second_report_df['OSM DIM Surcharge'] > 0]
osm_dim_count = osm_dim_df.shape[0]  # Count of non-null rows
osm_dim_sum = osm_dim_df['OSM DIM Surcharge'].sum()  # Sum

# Count and sum for Relabel Fee
relabel_fee_df = second_report_df[second_report_df['Relabel Fee'] > 0]
relabel_fee_count = relabel_fee_df.shape[0]
relabel_fee_sum = relabel_fee_df['Relabel Fee'].sum()

# Count and sum for Delivery Area Surcharge (DAS)
delivery_area_df = second_report_df[second_report_df['Delivery Area Surcharge (DAS)'] > 0]
delivery_area_count = delivery_area_df.shape[0]
delivery_area_sum = delivery_area_df['Delivery Area Surcharge (DAS)'].sum()

# Count and sum for OCR Fee
ocr_fee_df = second_report_df[second_report_df['OCR Fee'] > 0]
ocr_fee_count = ocr_fee_df.shape[0]
ocr_fee_sum = ocr_fee_df['OCR Fee'].sum()

# Count and sum for Peak Season Surcharge
peak_season_df = second_report_df[second_report_df['Peak Season Surcharge'] > 0]
peak_season_count = peak_season_df.shape[0]
peak_season_sum = peak_season_df['Peak Season Surcharge'].sum()

# Count and sum for Nonstandard Length Fee 22 in
nonstandard_length_22_df = second_report_df[second_report_df['Nonstandard Length Fee 22 in'] > 0]
nonstandard_length_22_count = nonstandard_length_22_df.shape[0]
nonstandard_length_22_sum = nonstandard_length_22_df['Nonstandard Length Fee 22 in'].sum()

# Count and sum for Nonstandard Length Fee 30 in
nonstandard_length_30_df = second_report_df[second_report_df['Nonstandard Length Fee 30 in'] > 0]
nonstandard_length_30_count = nonstandard_length_30_df.shape[0]
nonstandard_length_30_sum = nonstandard_length_30_df['Nonstandard Length Fee 30 in'].sum()

# Count and sum for Nonstandard Length Fee 2 cu
nonstandard_length_2cu_df = second_report_df[second_report_df['Nonstandard Length Fee 2 cu'] > 0]
nonstandard_length_2cu_count = nonstandard_length_2cu_df.shape[0]
nonstandard_length_2cu_sum = nonstandard_length_2cu_df['Nonstandard Length Fee 2 cu'].sum()

# Count and sum for Dimension Noncompliance
dimension_noncompliance_df = second_report_df[second_report_df['Dimension Noncompliance'] > 0]
dimension_noncompliance_count = dimension_noncompliance_df.shape[0]
dimension_noncompliance_sum = dimension_noncompliance_df['Dimension Noncompliance'].sum()

# Count and sum for Irregular Shape Charge
irregular_shape_df = second_report_df[second_report_df['Irregular Shape Charge'] > 0]
irregular_shape_count = irregular_shape_df.shape[0]
irregular_shape_sum = irregular_shape_df['Irregular Shape Charge'].sum()

# Count and sum for Non Compliance/Unmanifested Charge
non_compliance_df = second_report_df[second_report_df['Non Compliance/Unmanifested Charge'] > 0]
non_compliance_count = non_compliance_df.shape[0]
non_compliance_sum = non_compliance_df['Non Compliance/Unmanifested Charge'].sum()

# Count and sum for Signature Confirmation
signature_confirmation_df = second_report_df[second_report_df['Signature Confirmation'] > 0]
signature_confirmation_count = signature_confirmation_df.shape[0]
signature_confirmation_sum = signature_confirmation_df['Signature Confirmation'].sum()

# Load the invoice workbook and select the active sheet
wb = load_workbook(invoice_file_path)
ws = wb.active

# Update cell G4 by adding 1 to the numeric portion
current_value = ws['G4'].value

# Check if the current value is in the expected format
if isinstance(current_value, str) and current_value.startswith('CM'):
    # Extract the numeric portion and increment it
    number_part = int(current_value[2:])  # Get the number after 'CM'
    new_value = f"CM{number_part + 1}"     # Increment by 1 and reconstruct
    ws['G4'] = new_value                    # Update G4 with the new value
else:
    print("Unexpected format in G4: ", current_value)

# Update cells with the respective counts, sums, and static values
ws['F14'] = parcel_count
ws['G14'] = parcel_package_charge_sum
ws['F15'] = small_parcel_count
ws['G15'] = small_parcel_package_charge_sum
ws['F16'] = ground_adv_count
ws['G16'] = ground_adv_package_charge_sum
ws['F17'] = 1
ws['G17'] = fuel_charge_sum
ws['F18'] = delivery_confirmation_count
ws['G18'] = delivery_confirmation_sum
ws['F19'] = osm_dim_count
ws['G19'] = osm_dim_sum
ws['F20'] = relabel_fee_count
ws['G20'] = relabel_fee_sum
ws['F21'] = delivery_area_count
ws['G21'] = delivery_area_sum
ws['F22'] = ocr_fee_count
ws['G22'] = ocr_fee_sum
ws['F23'] = peak_season_count
ws['G23'] = peak_season_sum
ws['F24'] = nonstandard_length_22_count
ws['G24'] = nonstandard_length_22_sum
ws['F25'] = nonstandard_length_30_count
ws['G25'] = nonstandard_length_30_sum
ws['F26'] = nonstandard_length_2cu_count
ws['G26'] = nonstandard_length_2cu_sum
ws['F27'] = dimension_noncompliance_count
ws['G27'] = dimension_noncompliance_sum
ws['F28'] = irregular_shape_count
ws['G28'] = irregular_shape_sum
ws['F29'] = non_compliance_count
ws['G29'] = non_compliance_sum
ws['F30'] = signature_confirmation_count
ws['G30'] = signature_confirmation_sum


# Get today's date in the format yyyy-mm-dd
today_date = datetime.today().strftime('%Y-%m-%d')

# Update cell G3 with today's date
ws['G3'] = today_date

print(f"Today's date set in G3 as: {today_date}")


# Calculate the date for the most recent Saturday (yesterday if today is Sunday)
today = datetime.today()
last_saturday = today - timedelta(days=today.weekday() + 2)  # Saturday of the previous week
last_sunday = last_saturday - timedelta(days=6)  # Previous week's Sunday

# Format the date as MM/DD/YY for display in the cell
service_period = f"{last_sunday.strftime('%m/%d/%y')} - {last_saturday.strftime('%m/%d/%y')}"

# Update the merged cell (F9 and G9) with the calculated service period
ws['F9'] = service_period
print(f"Service period for the previous week set as: {service_period}")

# Unhide rows 13 to 30 before hiding specific rows
for row in range(13, 31):
    ws.row_dimensions[row].hidden = False

# Hide rows where the value in column G is empty for rows 13 to 30
for row in range(13, 31):  # Rows 13 to 30
    if ws[f'G{row}'].value == 0:
        ws.row_dimensions[row].hidden = True

# Save the modified invoice in the same folder as the input file
wb.save(modified_invoice_path)

print("Invoice updated and saved successfully.")
